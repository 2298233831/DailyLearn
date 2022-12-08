#coding=gbk
#��һ��ָ���˱����ʽ
from asyncio.windows_events import NULL  #����NULLֵ
from openpyxl import load_workbook  #����openpyxl�⣬�����ȡexcel�ļ�
from pathlib import Path    #����pathlib�⣬��������ļ�
import matplotlib.pyplot as plt #��ͼ��
import numpy as np  #numpy�⣬�����������
#��ʼ������
fpath=Path(input("Input the file path:"))#e.g.:D:\ҩƷ����1.xlsx
file_name=fpath.name    #��ȡ�ļ���
#������ֱ���
tmp=0
variance=0
print("Please input the number of data in the longest column :")
n=int(input())
data1=[0]*(n)  #����һ������Ϊn���б�,nΪ�����һ�еĳ���,���ʼ��ĩβ
data2=[0]*(n-1)   #�������ݣ����������ͼ,���ᱣ���е�����
No=[0]*(n-1)
for i in range(0,n-1):
    No[i]=i+1
    if i==n-1:
        break
#��ʼ��ȡexcel�ļ�
wb=load_workbook(file_name)  #��ȡexcel�ļ�
print(wb.sheetnames)    #��ȡexcel�ļ��е�sheet����
sheet=wb.active    #��ȡ��ǰ���sheet
print("Please input the scope of the data you want to read:")
column_start=int(input("column start:"))  #����ʼֵ���������������ݵ�һ����Ϊ��ʼֵ��
column_stop=int(input("column stop:")) 
row_start=int(input("raw start:")) #����ʼֵ
row_stop=int(input("raw stop:")) 
for i in range(column_start,column_stop+1): #��ȡÿ��ÿ�е�����
    for j in range(row_start,row_stop+1):   #��Ҫÿһ�еĿ�ͷ���������ƣ������б��û�������ưɡ��˴����벻�Ὣ�����Ƽ���������ݴ���
        cell1=sheet.cell(row=j,column=i)
        if j==1:
            num=NULL
            data1[j-1]=0
        else:
            num=cell1.value
            data1[j-1]=num
            data2[j-2]=data1[j-1]
            tmp+=num
        print(cell1.value)   #���ÿ����Ԫ�������
    num=0
    average_data=tmp/(row_stop-row_start)   #����ƽ��ֵ
    for k in range(row_start,row_stop+1):   #���㷽��
        num=data1[k-1]
        if num==0:
            variance=0   
        else:    
            variance+=(num-average_data)**2
    variance=variance/(row_stop-row_start-1)    #�����׼����
    print("The average data of this factory is:",'%.3f' %average_data)  #���ƽ��ֵ
    print("The variance of this factory is:",variance)  #�������
    tmp=0
    #��һ��ͼ--��ֱ��ɢ��ͼ��
    plt.figure(figsize=(10,8))
    plt.scatter(No,data2)
    plt.xlabel("No")
    plt.ylabel("data")
    plt.title("data & No")
    plt.show()
#����
print("Over!")